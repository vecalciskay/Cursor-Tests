"""Read sales data from Excel and write per-rep aggregates to JSON."""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
INPUT_PATH = REPO_ROOT / "inputs" / "data.xlsx"
OUTPUT_PATH = REPO_ROOT / "outputs" / "results.json"

EXPECTED_COLUMNS = (
    "Postcode",
    "Sales_Rep_ID",
    "Sales_Rep_Name",
    "Year",
    "Value",
)
REP_NAME_COLUMN = "Sales_Rep_Name"
VALUE_COLUMN = "Value"


def load_sales_rows(path: Path) -> list[dict[str, object]]:
    if not path.is_file():
        raise FileNotFoundError(f"Input file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None or sheet.max_row is None or sheet.max_row < 1:
            raise ValueError(f"Worksheet is empty in {path}")

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        column_index = _validate_headers(headers)

        rows: list[dict[str, object]] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            rep_name = row[column_index[REP_NAME_COLUMN]]
            value = row[column_index[VALUE_COLUMN]]

            if rep_name is None or str(rep_name).strip() == "":
                raise ValueError(
                    f"Missing {REP_NAME_COLUMN} at row {row_number} in {path}"
                )
            if value is None:
                raise ValueError(f"Missing {VALUE_COLUMN} at row {row_number} in {path}")
            if not isinstance(value, (int, float)):
                raise ValueError(
                    f"Invalid {VALUE_COLUMN} at row {row_number} in {path}: {value!r}"
                )

            rows.append(
                {
                    REP_NAME_COLUMN: str(rep_name).strip(),
                    VALUE_COLUMN: float(value),
                }
            )

        if not rows:
            raise ValueError(f"No data rows found in {path}")

        return rows
    finally:
        workbook.close()


def _validate_headers(headers: list[str]) -> dict[str, int]:
    missing = [name for name in EXPECTED_COLUMNS if name not in headers]
    if missing:
        raise ValueError(
            f"Missing expected column(s): {', '.join(missing)}. "
            f"Found: {', '.join(headers)}"
        )
    return {name: headers.index(name) for name in EXPECTED_COLUMNS}


def aggregate_by_rep(rows: list[dict[str, object]]) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        rep_name = str(row[REP_NAME_COLUMN])
        totals[rep_name] += float(row[VALUE_COLUMN])
    return dict(totals)


def build_result(aggregates: dict[str, float]) -> dict[str, list[dict[str, object]]]:
    # Sorted by salesRepName ascending for stable, predictable output.
    per_sale = [
        {"salesRepName": name, "totalValue": aggregates[name]}
        for name in sorted(aggregates)
    ]
    return {"perSale": per_sale}


def write_json(data: dict[str, object], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(data, file, indent=4)
        file.write("\n")


def main() -> int:
    rows = load_sales_rows(INPUT_PATH)
    aggregates = aggregate_by_rep(rows)
    result = build_result(aggregates)
    write_json(result, OUTPUT_PATH)
    print(f"Wrote {len(result['perSale'])} sales rep totals to {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, ValueError) as error:
        print(f"Error: {error}", file=sys.stderr)
        raise SystemExit(1) from error
